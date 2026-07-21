#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
BENCHMARK LUZ CNMC - VERSION PYTHON (equivalente a Benchmark.ps1)
- 28 empresas objetivo: se toman del API CNMC SOLO si su oferta es de PRECIO FIJO
  (tipoElectricidad='TE'); las que no aparecen asi se leen de su web.
- Casos especiales: Endesa y Energya VM -> energia CON descuento.
- Empresas forzadas siempre a web (su oferta CNMC no es la fija estandar):
  Octopus, Nexus, Iberdrola, El Corte Ingles, Fenie, Disa, Gana, Energy Asset.
- Excel con historico semanal (pestana por FECHA) + comparativa, en Descargas.
Requisitos: pandas/openpyxl/requests. Sin arrastre: si no hay dato -> 'No disponible'.
"""

import sys
# Instala dependencias solo si faltan (rapido; no fuerza upgrade cada vez)
try:
    import requests
    from openpyxl import load_workbook, Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "-q",
                           "openpyxl", "requests"])
    import requests
    from openpyxl import load_workbook, Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

import os
import re
import html
import json
import unicodedata
from datetime import datetime, timedelta
from pathlib import Path

import urllib3
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

# ------------------------------------------------------------------ CONFIG -----
CODIGO_POSTAL    = 28003
POTENCIA         = 4
CONSUMO_ANUAL_E  = 210
IEE              = 0.051    # Impuesto Especial sobre la Electricidad
IVA              = 0.21
TARIFA           = 4        # peaje 2.0TD domestico (NO 1: el API devuelve 0 ofertas)
REVISION_PRECIOS = 1        # 1 = precio fijo (CNMC)
CARPETA_SALIDA   = Path(os.environ.get('USERPROFILE', str(Path.home()))) / 'Downloads'
NOMBRE_EXCEL     = 'Analisis_Energia_CNMC.xlsx'

BASE_URL = 'https://comparador.cnmc.gob.es/api/publico'
HEADERS  = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'}

# NOMBRE_LEGAL (en el campo comercializadora) -> marca / cadena de busqueda / exclude
TARGET_OFFERS = {
    'ENERGYA VM':      {'brand': 'Energya VM',      'search': 'Formula Fija Unica'},
    'DOMESTICA GAS':   {'brand': 'Visalia',         'search': 'Plan Ahorro', 'exclude': '3 precios'},
    'NATURGY':         {'brand': 'Naturgy',         'search': 'Por Uso Luz'},
    'REPSOL':          {'brand': 'Repsol',          'search': 'Sin Horarios', 'exclude': 'mantenimiento|pyme|renovable|discrimina|veh'},
    'OCTOPUS':         {'brand': 'Octopus',         'search': 'OCTOPUS RELAX'},
    'NIBA':            {'brand': 'Niba',            'search': 'niba Zen'},
    'ENERGIA NUFRI':   {'brand': 'Energia Nufri',   'search': 'CALMA'},
    'GAOLANIA':        {'brand': 'Gana Energia',    'search': 'Tarifa 24 horas'},
    'CLEARVIEW':       {'brand': 'Clarity Energy',  'search': 'CLARITY', 'exclude': '3 precios|3p'},
    'CIDE':            {'brand': 'CHC Energia',     'search': 'Ilumina'},
    'ENERGYASSET':     {'brand': 'Energy Asset',    'search': 'Tarifa Plana'},
    'CATGAS':          {'brand': 'Catgas',          'search': '2.0TDL'},
    'TELECOR':         {'brand': 'El Corte Ingles', 'search': 'Despreocupate'},
    'ENDESA':          {'brand': 'Endesa',          'search': 'Luz Fija 24h'},
    'FENIE ENERGIA':   {'brand': 'Fenie Energia',   'search': 'Fijo Energetico'},
    'GESTERNOVA':      {'brand': 'Contigo Energia', 'search': 'Tarifa Facil'},
    'DISA ENERGIA':    {'brand': 'Disa Energia',    'search': 'ALISIOS'},
    'HIDROELECTRICA':  {'brand': 'HSC',             'search': 'Eficiente'},
    'LUMISA':          {'brand': 'Lumisa Energia',  'search': '2.0TD'},
    'TOTALENERGIES':   {'brand': 'Total Energies',  'search': 'TU AIRE'},
    'PLENITUDE':       {'brand': 'Plenitude',       'search': 'Tarifa Facil Plus'},
    'IMAGINA':         {'brand': 'Imagina',         'search': 'PLAN BASE'},
    'IBERDROLA':       {'brand': 'Iberdrola',       'search': 'Plan Online'},
    'NEXUS':           {'brand': 'Nexus',           'search': 'Luz Eficiente'},
    # Estas 3 se intentan primero en CNMC; si no estan, caen a web scraping
    'HOLALUZ':         {'brand': 'Holaluz',         'search': 'Tarifa Clasica'},
    'GEO ALTERNATIVA': {'brand': 'Podo',            'search': 'Luz Precio Unico 24h'},
    'FACTORENERGIA':   {'brand': 'Factor Energia',  'search': 'Tarifa Unica'},
}

# --------------------------------------------------------------- HELPERS -------

def build_params(id_oferta=0):
    today = datetime.now()
    fin = today.replace(day=1) - timedelta(days=1)   # ultimo dia del mes anterior
    inicio = fin.replace(day=1)                       # primer dia del mes anterior
    ts_ini = int(inicio.timestamp() * 1000)
    ts_fin = int(fin.timestamp() * 1000)
    return {
        'tipoSuministro': 'E', 'codigoPostal': CODIGO_POSTAL, 'potencia': POTENCIA,
        'potenciaPrimeraFranja': POTENCIA, 'potenciaSegundaFranja': POTENCIA,
        'potenciaTerceraFranja': POTENCIA, 'potenciaCuartaFranja': POTENCIA,
        'potenciaQuintaFranja': POTENCIA, 'potenciaSextaFranja': POTENCIA,
        'consumoAnualE': CONSUMO_ANUAL_E, 'consumoAnualEOrig': 2600,
        'consumoPrimeraFranja': 61, 'consumoSegundaFranja': 51,
        'consumoTerceraFranja': 98, 'consumoCuartaFranja': 0,
        'consumoQuintaFranja': 0, 'consumoSextaFranja': 0,
        'consumoAnualEQr': 0, 'consumoPrimeraFranjaQr': 0,
        'consumoSegundaFranjaQr': 0, 'consumoTerceraFranjaQr': 0,
        'consumoCuartaFranjaQr': 0, 'consumoQuintaFranjaQr': 0,
        'consumoSextaFranjaQr': 0, 'consumoAnualEPQr': 0,
        'consumoPrimeraFranjaPQr': 0, 'consumoSegundaFranjaPQr': 0,
        'consumoTerceraFranjaPQr': 0, 'consumoCuartaFranjaPQr': 0,
        'consumoQuintaFranjaPQr': 0, 'consumoSextaFranjaPQr': 0,
        'tarifa': TARIFA, 'consumoAnualG': 491, 'consumoAnualGOrig': 6000,
        'serviciosAdicionales': 2, 'permanencia': 2, 'idOferta': id_oferta,
        'vivienda': 'true', 'factura': 'true',
        'energiaAutoconsumo': 0, 'idAuditoriaQR': 0,
        'potenciaAutoconsumo': 3.5, 'revisionPrecios': REVISION_PRECIOS, 'importe': 0,
        'dateInicio': ts_ini, 'dateFin': ts_fin,
        'tc': 0, 'bs': 0, 'impSA': 0, 'impOtros': 0, 'exc': 0, 'reg': 0,
        'mecanismoAjuste': 0, 'importeMecanismoAjustePunta': 0,
        'importeMecanismoAjusteLlano': 0, 'importeMecanismoAjusteValle': 0,
        'precioConsumoMecanismoAjustePunta': 0, 'precioConsumoMecanismoAjusteLlano': 0,
        'precioConsumoMecanismoAjusteValle': 0, 'precioConsumoMecanismoAjusteTotal': 0,
        'mecanismoAjusteIVA': 0, 'impOtrosConIE': 0, 'impOtrosSinIE': 0,
        'pmaxP1': 0, 'pmaxP2': 0, 'fFact': ts_fin,
        'dtoBS': 0, 'finBS': 0, 'ajuste': 0, 'impPot': 0, 'impEner': 0, 'dto': 0,
        'prP1': 0, 'prP2': 0, 'prE1': 0, 'prE2': 0, 'prE3': 0,
        'cfP1flex': 0, 'cfP2flex': 0, 'cambio': 0, 'promo': 0, 'verde': 0,
        'rev': 0, 'trampeo': 0,
        'perfilConsumo': 13, 'cups': '0000', 'autoconsumo': 'false',
    }


def cnmc_get(path, extra=None):
    """GET al API CNMC, decodificando SIEMPRE como UTF-8 (evita ENERGIA mal codificada)."""
    params = build_params()
    if extra:
        params.update(extra)
    r = requests.get(f'{BASE_URL}{path}', params=params, headers=HEADERS,
                     timeout=60, verify=False)
    return json.loads(r.content.decode('utf-8'))


def get_web_text(url):
    """Descarga y decodifica entidades HTML (&euro; -> simbolo, &iacute; -> i, ...)."""
    try:
        r = requests.get(url, headers=HEADERS, timeout=20, verify=False)
        r.encoding = 'utf-8'
        return html.unescape(r.text)
    except Exception:
        return None


def strip_tags(text):
    return re.sub(r'\s+', ' ', re.sub(r'<[^>]+>', ' ', text))


def get_number(text):
    if text is None or text == '':
        return None
    try:
        return float(str(text).replace(',', '.'))
    except ValueError:
        return None


def find_in_range(zona, patrones, mn, mx):
    """Primer numero (segun patrones) que caiga en [mn, mx]."""
    for p in patrones:
        for m in re.finditer(p, zona, re.IGNORECASE):
            v = get_number(m.group(1))
            if v is not None and mn <= v <= mx:
                return v
    return None


def parse_prices(texto):
    """
    Devuelve dict energia/p1/p2 a partir del texto 'caracteristicas'.
    Estrategia por SECCIONES: separa la zona de POTENCIA y la de ENERGIA por sus
    encabezados (el orden varia por comercializadora) y busca dentro de cada una.
    '.' en los patrones sustituye a letras acentuadas (e/i/n).
    """
    r = {'energia': None, 'p1': None, 'p2': None}
    if not texto:
        return r
    t = texto.replace('\xa0', ' ')

    re_hdr_pot = r'(?:precios?\s+(?:del\s+)?(?:t.rmino\s+)?(?:de\s+)?potencia|t.rmino[s]?\s+(?:de\s+|de\s+la\s+)?potencia)(?=[\s\S]{0,32}[0-9]+[.,][0-9])'
    re_hdr_ene = r'(?:precios?\s+(?:del\s+)?(?:t.rmino\s+)?(?:de\s+)?energ.a|t.rmino[s]?\s+(?:de\s+|de\s+la\s+)?energ.a)(?=[\s\S]{0,32}[0-9]+[.,][0-9])'
    m_pot = re.search(re_hdr_pot, t, re.IGNORECASE)
    m_ene = re.search(re_hdr_ene, t, re.IGNORECASE)

    zona_pot = t
    zona_ene = t
    if m_pot and m_ene:
        if m_pot.start() < m_ene.start():
            zona_pot = t[m_pot.start():m_ene.start()]
            zona_ene = t[m_ene.start():]
        else:
            zona_ene = t[m_ene.start():m_pot.start()]
            zona_pot = t[m_pot.start():]
    elif m_pot:
        zona_pot = t[m_pot.start():]
    elif m_ene:
        zona_ene = t[m_ene.start():]

    # Energia (E/kWh) 0.05-0.30: preferir precio SIN descuento
    r['energia'] = find_in_range(zona_ene, [
        r'sin\s+descuentos?.*?([0-9]+[.,][0-9]{3,6})',
        r'([0-9]+[.,][0-9]{3,6})\s*[^0-9]{0,4}kWh',
        r'([0-9]+[.,][0-9]{3,6})',
    ], 0.05, 0.30)

    # Potencia P1 (E/kW ano) 0-75
    r['p1'] = find_in_range(zona_pot, [
        r'(?:^|[^A-Za-z])P1\b.*?([0-9]+[.,][0-9]+)',
        r'[Pp]unta.*?([0-9]+[.,][0-9]+)',
        r'[Pp]eriodo\s+1.*?([0-9]+[.,][0-9]+)',
    ], 0, 75)

    # Potencia P2 (E/kW ano) 0-75
    r['p2'] = find_in_range(zona_pot, [
        r'(?:^|[^A-Za-z])P2\b.*?([0-9]+[.,][0-9]+)',
        r'[Vv]alle.*?([0-9]+[.,][0-9]+)',
        r'(?:^|[^A-Za-z])P3\b.*?([0-9]+[.,][0-9]+)',
        r'[Pp]eriodo\s+2.*?([0-9]+[.,][0-9]+)',
    ], 0, 75)

    if r['p1'] is not None and r['p2'] is None:
        r['p2'] = r['p1']
    return r


# --------------------------------------------------------- WEB SCRAPERS --------

def _none():
    return {'energia': None, 'p1': None, 'p2': None}


def get_holaluz():
    text = get_web_text('https://www.holaluz.com/luz/')
    if not text:
        return _none()
    t = strip_tags(text)
    m = re.search(r'Tarifa\s+Cl.sica.{0,120}?Precio\s+24\s+horas\s+([0-9][.,][0-9]{2,6}).{0,40}?P1\s+([0-9][.,][0-9]{2,6}).{0,30}?P2\s+([0-9][.,][0-9]{2,6})', t, re.IGNORECASE)
    if m:
        e = get_number(m.group(1))
        p1 = round(get_number(m.group(2)) * 365, 2)
        p2 = round(get_number(m.group(3)) * 365, 2)
        if 0.05 <= e <= 0.30:
            return {'energia': e, 'p1': p1, 'p2': p2}
    e = find_in_range(t, [r'([0-9][.,][0-9]{2,6})\s*[^0-9]{0,6}kWh'], 0.05, 0.30)
    pd = find_in_range(t, [r'([0-9][.,][0-9]{2,6})\s*[^0-9]{0,6}kW(?!h|/kWh)'], 0.01, 0.60)
    if e is not None and pd is not None:
        p = round(pd * 365, 2)
        return {'energia': e, 'p1': p, 'p2': p}
    return _none()


def get_podo():
    text = get_web_text('https://www.mipodo.com/tarifas-luz')
    if not text:
        return _none()
    t = strip_tags(text)
    m = re.search(r'Energ.a\s*24h\s*([0-9][.,][0-9]{2,6})\s*[^0-9]{0,6}kWh\s*Potencia\s*P1\s*([0-9][.,][0-9]{2,6})\s*[^0-9]{0,10}kW.{0,4}d.a\s*P2\s*([0-9][.,][0-9]{2,6})', t, re.IGNORECASE)
    if m:
        e = get_number(m.group(1))
        p1 = round(get_number(m.group(2)) * 365, 2)
        p2 = round(get_number(m.group(3)) * 365, 2)
        if 0.05 <= e <= 0.30:
            return {'energia': e, 'p1': p1, 'p2': p2}
    return _none()


def get_factor():
    text = get_web_text('https://www.factorenergia.com/es/luz/tarifa-fija-de-luz-precio-unico/')
    if not text:
        return _none()
    m_e = re.search(r'(0[,\.]\d{3,4})\s*.{0,3}/kWh', text)
    m_p = re.search(r'([\d,\.]+)\s*.{0,3}/kW\s*(?:d.a|day)', text, re.IGNORECASE)
    if m_e and m_p:
        e = get_number(m_e.group(1))
        p = round(get_number(m_p.group(1)) * 365, 2)
        return {'energia': e, 'p1': p, 'p2': p}
    return _none()


def get_iberdrola():
    text = get_web_text('https://www.iberdrola.es/luz/plan-online')
    if not text:
        return _none()
    t = strip_tags(text)
    m_e = re.search(r'24\s*horas\s*del\s*d.a\s*([0-9][.,][0-9]{3,6})\s*[^0-9]{0,4}kWh', t, re.IGNORECASE)
    m_p1 = re.search(r'Periodo\s*Punta\s*([0-9][.,][0-9]{3,6})\s*[^0-9]{0,6}kW', t, re.IGNORECASE)
    m_p2 = re.search(r'Periodo\s*Valle\s*([0-9][.,][0-9]{3,6})\s*[^0-9]{0,6}kW', t, re.IGNORECASE)
    if m_e and m_p1 and m_p2:
        e = get_number(m_e.group(1))
        p1 = round(get_number(m_p1.group(1)) * 365, 2)
        p2 = round(get_number(m_p2.group(1)) * 365, 2)
        if 0.05 <= e <= 0.30:
            return {'energia': e, 'p1': p1, 'p2': p2}
    return _none()


def get_total_energies():
    text = get_web_text('https://www.totalenergies.es/es/hogares/tarifas-luz/a-tu-aire-siempre')
    if not text:
        return _none()
    t = strip_tags(text)
    m_e = re.search(r'precio\s+fijo\s+a\s*([0-9][.,][0-9]{2,6})\s*[^0-9]{0,5}kWh', t, re.IGNORECASE)
    m_p1 = re.search(r'Potencia\s*P1\s*([0-9][.,][0-9]{3,6})\s*[^0-9]{0,5}kW', t, re.IGNORECASE)
    m_p2 = re.search(r'\bP2\s*([0-9][.,][0-9]{3,6})\s*[^0-9]{0,5}kW', t, re.IGNORECASE)
    if m_e and m_p1:
        e = get_number(m_e.group(1))
        p1 = round(get_number(m_p1.group(1)) * 365, 2)
        p2 = round(get_number(m_p2.group(1)) * 365, 2) if m_p2 else p1
        if 0.05 <= e <= 0.30:
            return {'energia': e, 'p1': p1, 'p2': p2}
    return _none()


def get_plenitude():
    text = get_web_text('https://eniplenitude.es/hogar/tarifas-luz/facil/')
    if not text:
        return _none()
    t = strip_tags(text)
    m_e = re.search(r'Precio\s+especial:\s*([0-9][.,][0-9]{2,6})\s*[^0-9]{0,5}kWh', t, re.IGNORECASE)
    m_p = re.search(r'POTENCIA.{0,80}?Sin\s+impuestos\s*([0-9][.,][0-9]{3,6})\s+([0-9][.,][0-9]{3,6})', t, re.IGNORECASE)
    if m_e and m_p:
        e = get_number(m_e.group(1))
        p1 = round(get_number(m_p.group(1)) * 365, 2)
        p2 = round(get_number(m_p.group(2)) * 365, 2)
        if 0.05 <= e <= 0.30:
            return {'energia': e, 'p1': p1, 'p2': p2}
    return _none()


def get_octopus():
    text = get_web_text('https://octopusenergy.es/precios')
    if not text:
        return _none()
    t = strip_tags(text)
    m = re.search(r'([0-9][.,][0-9]{2,6})\s*[^0-9]{0,5}kWh\s*10%\s*dto.{0,80}?Potencia\s*Punta\s*\(P1\)\s*([0-9][.,][0-9]{2,6}).{0,40}?Valle\s*\(P2\)\s*([0-9][.,][0-9]{2,6})', t, re.IGNORECASE)
    if m:
        e = get_number(m.group(1))
        p1 = round(get_number(m.group(2)) * 365, 2)
        p2 = round(get_number(m.group(3)) * 365, 2)
        if 0.05 <= e <= 0.30:
            return {'energia': e, 'p1': p1, 'p2': p2}
    return _none()


def get_clarity():
    text = get_web_text('https://www.clarityenergy.es/tarifas-luz/')
    if not text:
        return _none()
    t = strip_tags(text)
    m_e = re.search(r'Precio\s+Energ.a\s+24\s+Horas\s*:\s*([0-9][.,][0-9]{2,6})', t, re.IGNORECASE)
    m_p = re.search(r'Precios?\s+Potencia\s+P1\s*:\s*([0-9][.,][0-9]{2,6})\s*[^0-9]{0,6}kW.{0,25}?P2\s*:\s*([0-9][.,][0-9]{2,6})', t, re.IGNORECASE)
    if m_e and m_p:
        e = get_number(m_e.group(1))
        p1 = round(get_number(m_p.group(1)) * 365, 2)
        p2 = round(get_number(m_p.group(2)) * 365, 2)
        if 0.05 <= e <= 0.30:
            return {'energia': e, 'p1': p1, 'p2': p2}
    return _none()


def get_nexus():
    # Tarifa plana: PUNTA=LLANO=VALLE en energia (backref \1). Potencia EUR/kW MES -> x12.
    text = get_web_text('https://www.nexusenergia.com/hogar/tarifas-luz/')
    if not text:
        return _none()
    t = strip_tags(text)
    m = re.search(r'Precio\s+energ.a\s+PUNTA\s+([0-9][.,][0-9]{2,6}).{0,80}?LLANO\s+\1.{0,80}?VALLE\s+\1.{0,140}?Precio\s+potencia\s+PUNTA\s+([0-9][.,][0-9]{2,6}).{0,90}?VALLE\s+([0-9][.,][0-9]{2,6})', t, re.IGNORECASE)
    if m:
        e = get_number(m.group(1))
        p1 = round(get_number(m.group(2)) * 12, 2)
        p2 = round(get_number(m.group(3)) * 12, 2)
        if 0.05 <= e <= 0.30:
            return {'energia': e, 'p1': p1, 'p2': p2}
    return _none()


def get_el_corte_ingles():
    text = get_web_text('https://www.energiaelcorteingles.es/tarifa-despreocupate/')
    if not text:
        return _none()
    t = strip_tags(text)
    m_e = re.search(r'Periodo\s+las\s+24\s*h\.?\s*([0-9][.,][0-9]{3,6})', t, re.IGNORECASE)
    m_p1 = re.search(r'Tramo\s+Punta\s*([0-9][.,][0-9]{3,6})\s*[^0-9]{0,5}kW', t, re.IGNORECASE)
    m_p2 = re.search(r'Tramo\s+Valle\s*([0-9][.,][0-9]{3,6})\s*[^0-9]{0,5}kW', t, re.IGNORECASE)
    if m_e and m_p1:
        e = get_number(m_e.group(1))
        p1 = round(get_number(m_p1.group(1)) * 365, 2)
        p2 = round(get_number(m_p2.group(1)) * 365, 2) if m_p2 else p1
        if 0.05 <= e <= 0.30:
            return {'energia': e, 'p1': p1, 'p2': p2}
    return _none()


def get_fenie():
    text = get_web_text('https://www.fenieenergia.es/es/hogar/tarifas-de-luz/fijo-energetico-1p')
    if not text:
        return _none()
    t = strip_tags(text)
    m_e = re.search(r'T.rmino\s+de\s+energ.a\s+P1\s*([0-9][.,][0-9]{3,6})', t, re.IGNORECASE)
    m_p1 = re.search(r'T.rmino\s+de\s+potencia\s+P1\s*([0-9][.,][0-9]{3,6})', t, re.IGNORECASE)
    m_p2 = re.search(r'T.rmino\s+de\s+potencia\s+P1.{0,60}?P2\s*([0-9][.,][0-9]{3,6})', t, re.IGNORECASE)
    if m_e and m_p1:
        e = get_number(m_e.group(1))
        p1 = round(get_number(m_p1.group(1)) * 365, 2)
        p2 = round(get_number(m_p2.group(1)) * 365, 2) if m_p2 else p1
        if 0.05 <= e <= 0.30:
            return {'energia': e, 'p1': p1, 'p2': p2}
    return _none()


def get_disa():
    # Tabla: "SIN DTO <potP1> <potP2> <enP1> <enP2> <enP3>". Potencia EUR/kW dia -> x365.
    text = get_web_text('https://www.disagrupo.es/electricidad/tarifa-disa-alisios/')
    if not text:
        return _none()
    t = strip_tags(text)
    m = re.search(r'SIN\s*DTO\s*([0-9][.,][0-9]{3,6})\s+([0-9][.,][0-9]{3,6})\s+([0-9][.,][0-9]{3,6})', t, re.IGNORECASE)
    if m:
        p1 = round(get_number(m.group(1)) * 365, 2)
        p2 = round(get_number(m.group(2)) * 365, 2)
        e = get_number(m.group(3))
        if 0.05 <= e <= 0.30:
            return {'energia': e, 'p1': p1, 'p2': p2}
    return _none()


def get_chc():
    # "Plan Ilumina Duo Promo" (tarifa 24h de hogar). El nombre del plan aparece primero en un
    # menu de navegacion (sin precios cerca) y despues en su tarjeta real; se ancla a "...Promo
    # Viviendas" para coger la tarjeta con precios. EUR/kW/dia -> x365.
    text = get_web_text('https://chcenergia.es/')
    if not text:
        return _none()
    t = strip_tags(text)
    m = re.search(r'Plan\s+Ilumina\s+D.o\s+Promo\s+Viviendas[\s\S]{0,650}', t, re.IGNORECASE)
    if not m:
        return _none()
    blk = m.group(0)
    mE = re.search(r'Precio\s+([0-9][.,][0-9]{3,6})\s*€/kWh', blk)
    mP1 = re.search(r'Punta\s+([0-9][.,][0-9]{3,6})\s*€/kW', blk)
    mP2 = re.search(r'Valle\s+([0-9][.,][0-9]{3,6})\s*€/kW', blk)
    if mE:
        e = get_number(mE.group(1))
        p1 = round(get_number(mP1.group(1)) * 365, 2) if mP1 else None
        p2 = round(get_number(mP2.group(1)) * 365, 2) if mP2 else None
        if 0.05 <= e <= 0.30:
            return {'energia': e, 'p1': p1, 'p2': p2}
    return _none()


def get_gana():
    # Energia en JSON-LD ("price":"0.119"). La potencia no se publica en HTML (solo JS).
    text = get_web_text('https://ganaenergia.com/tarifas-luz/24-horas')
    if not text:
        return _none()
    m = re.search(r'price[\\"\s:]{1,10}([0-9]\.[0-9]{2,6})', text, re.IGNORECASE)
    if m:
        e = get_number(m.group(1))
        if 0.05 <= e <= 0.30:
            return {'energia': e, 'p1': None, 'p2': None}
    return _none()


def get_energy_asset():
    # SPA: precios en el bundle JS (/assets/index-*.js), objeto "Tarifa Plana.Lanzamiento".
    html_page = get_web_text('https://energyasset.es/tarifa')
    if not html_page:
        return _none()
    m_js = re.search(r'src="(/assets/index-[^"]+\.js)"', html_page)
    if not m_js:
        return _none()
    js = get_web_text('https://energyasset.es' + m_js.group(1))
    if not js:
        return _none()
    m = re.search(r'Tarifa Plana\.Lanzamiento[\s\S]{0,300}?energyPrices:\[([0-9\.]+)\][\s\S]{0,60}?powerPrices:\[([0-9\.]+),\s*([0-9\.]+)\]', js)
    if m:
        e = get_number(m.group(1))
        p1 = round(get_number(m.group(2)) * 365, 2)
        p2 = round(get_number(m.group(3)) * 365, 2)
        if 0.05 <= e <= 0.30:
            return {'energia': e, 'p1': p1, 'p2': p2}
    return _none()


def get_energyavm():
    # Energia: "Precio unico 24 horas <sinDto> <conDto>" -> con descuento (segundo).
    # Potencia: "P1(Punta): X P2(Valle): Y" en EUR/kW-ano (directo, sin x365).
    text = get_web_text('https://www.energyavm.es/luz/formula-fija-24-horas-luz/')
    if not text:
        return _none()
    t = strip_tags(text)
    e = None
    mE = re.search(r'Precio\s+.?nico\s+24\s+horas\s+([0-9][.,][0-9]{3,6})\s+([0-9][.,][0-9]{3,6})', t, re.IGNORECASE)
    if mE:
        e = get_number(mE.group(2))
    p1 = p2 = None
    mP = re.search(r'P1\(Punta\):\s*([0-9]{1,3}[.,][0-9]{2,6}).{0,30}?P2\(Valle\):\s*([0-9]{1,3}[.,][0-9]{2,6})', t, re.IGNORECASE)
    if mP:
        p1 = round(get_number(mP.group(1)), 2)
        p2 = round(get_number(mP.group(2)), 2)
    if e is not None and 0.05 <= e <= 0.30:
        return {'energia': e, 'p1': p1, 'p2': p2}
    return _none()


def get_visalia():
    # La web trae un JSON con varias tarifas. Se coge la de "Plan Ahorro 3 meses" del bloque
    # "products" (SIN impuestos): energia_p1 EUR/kWh + potencia_p1/p2 EUR/kW dia -> x365.
    text = get_web_text('https://visalia.es/luz/fijo24horas/')
    if not text:
        return _none()
    m = re.search(r'"name":"Plan Ahorro 3 meses".{0,80}?"products":\{([^}]+)\}', text, re.S)
    if not m:
        return _none()
    blk = m.group(1)

    def g(key):
        mm = re.search(r'"%s":([0-9.]+)' % key, blk)
        return get_number(mm.group(1)) if mm else None

    e = g('energia_p1')
    pp1, pp2 = g('potencia_p1'), g('potencia_p2')
    p1 = round(pp1 * 365, 2) if pp1 is not None else None
    p2 = round(pp2 * 365, 2) if pp2 is not None else None
    if e is not None and 0.05 <= e <= 0.30:
        return {'energia': e, 'p1': p1, 'p2': p2}
    return _none()


def get_niba():
    # "Precio 24h X EUR/kWh ... Potencia Valle Y EUR/kW dia Punta Z EUR/kW dia". P1=Punta, P2=Valle. x365.
    text = get_web_text('https://niba.es/luz-y-gas')
    if not text:
        return _none()
    t = strip_tags(text)
    mE = re.search(r'Precio\s*24h\s*([0-9][.,][0-9]{2,6})\s*[^0-9]{0,4}kWh', t, re.IGNORECASE)
    mP = re.search(r'Potencia\s*Valle\s*([0-9][.,][0-9]{2,6})\s*[^0-9]{0,8}kW.{0,6}d.a\s*Punta\s*([0-9][.,][0-9]{2,6})', t, re.IGNORECASE)
    if mE:
        e = get_number(mE.group(1))
        p1 = round(get_number(mP.group(2)) * 365, 2) if mP else None
        p2 = round(get_number(mP.group(1)) * 365, 2) if mP else None
        if 0.05 <= e <= 0.30:
            return {'energia': e, 'p1': p1, 'p2': p2}
    return _none()


def get_naturgy():
    # "X EUR/kWh Termino de potencia ... P1 Y EUR/kW*dia P2 Z EUR/kW*dia". x365.
    text = get_web_text('https://www.naturgy.es/hogar/luz')
    if not text:
        return _none()
    t = strip_tags(text)
    # Energia SIN impuestos = primer valor tras "Termino de energia" (el 2o es con IVA).
    mE = re.search(r'T.rmino\s+de\s+energ.a\s*([0-9][.,][0-9]{3,6})\s*[^0-9]{0,3}kWh', t, re.IGNORECASE)
    mP = re.search(r'T.rmino\s+de\s+potencia.{0,40}?P1\s*([0-9][.,][0-9]{3,6})\s*[^0-9]{0,6}kW.{0,6}d.a\s*P2\s*([0-9][.,][0-9]{3,6})', t, re.IGNORECASE)
    if mE:
        e = get_number(mE.group(1))
        p1 = round(get_number(mP.group(1)) * 365, 2) if mP else None
        p2 = round(get_number(mP.group(2)) * 365, 2) if mP else None
        if 0.05 <= e <= 0.30:
            return {'energia': e, 'p1': p1, 'p2': p2}
    return _none()


def get_catgas():
    # Tarifa "2.0 TDL": potencia anual directa "Punta P1 X EUR/kW ... Valle P2 Y EUR/kW" +
    # energia fija "Coste energia Lineal Z EUR/kWh" (la pagina tambien ofrece una tarifa
    # INDEXADA aparte, con el texto literal "Indexado" en vez de un numero; esa se ignora).
    text = get_web_text('https://catgas.cat/es/tarifes-catllum-llar')
    if not text:
        return _none()
    t = strip_tags(text)
    mP1 = re.search(r'Punta\s*P1\s*([0-9]{1,2}[.,][0-9]{1,4})\s*[^0-9]{0,3}kW', t, re.IGNORECASE)
    mP2 = re.search(r'Valle\s*P2\s*([0-9]{1,2}[.,][0-9]{1,4})\s*[^0-9]{0,3}kW', t, re.IGNORECASE)
    mE = re.search(r'Lineal\s*([0-9][.,][0-9]{3,6})\s*[^0-9]{0,3}kWh', t, re.IGNORECASE)
    p1 = round(get_number(mP1.group(1)), 2) if mP1 else None
    p2 = round(get_number(mP2.group(1)), 2) if mP2 else None
    e = get_number(mE.group(1)) if mE else None
    if e is not None and not (0.05 <= e <= 0.30):
        e = None
    if e is not None or p1 is not None or p2 is not None:
        return {'energia': e, 'p1': p1, 'p2': p2}
    return _none()


def get_imagina():
    # "Ahora XX%: X EUR/kWh ... Potencia Valle: Y EUR/kW Punta: Z EUR/kW" (EUR/kW y dia -> x365).
    text = get_web_text('https://imaginaenergia.com/tarifa-luz-sin-horas/')
    if not text:
        return _none()
    t = strip_tags(text)
    mE = re.search(r'Ahora\s*\d{1,3}\s*%\s*:?\s*([0-9][.,][0-9]{2,6})', t, re.IGNORECASE)
    if not mE:
        mE = re.search(r'([0-9][.,][0-9]{2,6})\s*[^0-9]{0,4}kWh', t)
    mP = re.search(r'Valle:\s*([0-9][.,][0-9]{2,6})\s*[^0-9]{0,4}kW\s*Punta:\s*([0-9][.,][0-9]{2,6})', t, re.IGNORECASE)
    if mE:
        e = get_number(mE.group(1))
        p1 = round(get_number(mP.group(2)) * 365, 2) if mP else None
        p2 = round(get_number(mP.group(1)) * 365, 2) if mP else None
        if 0.05 <= e <= 0.30:
            return {'energia': e, 'p1': p1, 'p2': p2}
    return _none()


def get_nufri():
    # Tarifa plana "consume sin mirar el reloj" (24h, precio fijo 12 meses).
    # Potencia: viene en el JSON embebido (comillas escapadas \" al estar anidado en un string JS).
    # Se toma el bloque powerPrice{p1,p2} que va justo antes del uniqueEnergyPrice que coincide
    # con la energia ya extraida (asi se coge la potencia del producto correcto). EUR/kW/dia -> x365.
    text = get_web_text('https://www.energianufri.com/es/tarifas-luz')
    if not text:
        return _none()
    t = strip_tags(text)
    m = re.search(r'sin\s+mirar\s+el\s+reloj.{0,90}?([0-9][.,][0-9]{2,4})\s*[^0-9]{0,6}kWh', t, re.IGNORECASE)
    if not m:
        return _none()
    e = get_number(m.group(1))
    if not (0.05 <= e <= 0.30):
        return _none()
    p1 = p2 = None
    pat = r'powerPrice\\?"\s*:\s*\{\\?"p1\\?"\s*:\s*([0-9.]+)\s*,\s*\\?"p2\\?"\s*:\s*([0-9.]+)[^}]*\}\s*,\s*\\?"energyPrice\\?"\s*:\s*\{[^}]*\}\s*,\s*\\?"uniqueEnergyPrice\\?"\s*:\s*([0-9.]+)'
    for pm in re.finditer(pat, text):
        if abs(get_number(pm.group(3)) - e) < 0.0005:
            p1 = round(get_number(pm.group(1)) * 365, 2)
            p2 = round(get_number(pm.group(2)) * 365, 2)
            break
    return {'energia': e, 'p1': p1, 'p2': p2}


def get_lumisa():
    # La pagina "tarifa-fija" incluye TAMBIEN el plan "Lumisa Simple 24h" (precio unico):
    # "Precio Energia Las 24 horas : X EUR/kWh". La potencia (Precio Potencia P1/P2 EUR/kW/dia)
    # es la misma para todos los planes de la web -> x365.
    text = get_web_text('https://lumisa.es/es/tarifa-fija')
    if not text:
        return _none()
    t = strip_tags(text)
    mE = re.search(r'Precio\s+Energ.a\s+Las\s+24\s+horas\s*:\s*([0-9][.,][0-9]{3,6})\s*[^0-9]{0,4}kWh', t, re.IGNORECASE)
    mP = re.search(r'Precio\s+Potencia\s+P1\s*:\s*([0-9][.,][0-9]{3,6})\s*[^0-9]{0,6}kW.{0,6}d.a\s*P2\s*:\s*([0-9][.,][0-9]{3,6})', t, re.IGNORECASE)
    e = get_number(mE.group(1)) if mE else None
    if e is not None and not (0.05 <= e <= 0.30):
        e = None
    p1 = round(get_number(mP.group(1)) * 365, 2) if mP else None
    p2 = round(get_number(mP.group(2)) * 365, 2) if mP else None
    if e is not None or p1 is not None:
        return {'energia': e, 'p1': p1, 'p2': p2}
    return _none()


def get_endesa():
    # Energia: "...kWh que consumas al mes. <conDto> <sinDto> EUR/kWh EUR/kWh" -> se coge conDto
    # (mismo caso especial que en CNMC). Cuidado: la web tiene un numero de ejemplo en un FAQ
    # ("T. Energia: 0,163097 EUR/kWh") que NO es el precio real, hay que ignorarlo.
    # Potencia: "T. potencia hora valle ... X EUR/kW" y "... punta-llano ... X EUR/kW" (EUR/kW/mes) -> x12.
    text = get_web_text('https://www.endesa.com/es/luz-y-gas/luz/one/tarifa-one-luz')
    if not text:
        return _none()
    t = strip_tags(text)
    mE = re.search(r'kWh\s+que\s+consumas\s+al\s+mes\.\s*([0-9][.,][0-9]{3,6})\s+[0-9][.,][0-9]{3,6}\s*.{0,3}kWh', t, re.IGNORECASE)
    # anclado a "Es el precio por kW..." para no caer en el ejemplo de la FAQ (mismo texto "potencia hora valle/punta-llano" pero con otro precio de relleno).
    mPv = re.search(r'potencia\s+hora\s+valle\s+Es\s+el\s+precio\s+por\s+kW.{0,320}?([0-9][.,][0-9]{3,6})\s*.{0,2}kW(?!h)', t, re.IGNORECASE)
    mPp = re.search(r'potencia\s+hora\s+punta-llano\s+Es\s+el\s+precio\s+por\s+kW.{0,320}?([0-9][.,][0-9]{3,6})\s*.{0,2}kW(?!h)', t, re.IGNORECASE)
    if mE:
        e = get_number(mE.group(1))
        p1 = round(get_number(mPp.group(1)) * 12, 2) if mPp else None
        p2 = round(get_number(mPv.group(1)) * 12, 2) if mPv else None
        if 0.05 <= e <= 0.30:
            return {'energia': e, 'p1': p1, 'p2': p2}
    return _none()


SCRAPERS = {
    'Energya VM':      get_energyavm,
    'Endesa':          get_endesa,
    'Visalia':         get_visalia,
    'Niba':            get_niba,
    'Naturgy':         get_naturgy,
    'Catgas':          get_catgas,
    'Imagina':         get_imagina,
    'Energia Nufri':   get_nufri,
    'Lumisa Energia':  get_lumisa,
    'Holaluz':         get_holaluz,
    'Podo':            get_podo,
    'Factor Energia':  get_factor,
    'Iberdrola':       get_iberdrola,
    'Total Energies':  get_total_energies,
    'Plenitude':       get_plenitude,
    'Octopus':         get_octopus,
    'Clarity Energy':  get_clarity,
    'Nexus':           get_nexus,
    'El Corte Ingles': get_el_corte_ingles,
    'Fenie Energia':   get_fenie,
    'Disa Energia':    get_disa,
    'CHC Energia':     get_chc,
    'Gana Energia':    get_gana,
    'Energy Asset':    get_energy_asset,
}


# --------------------------------------------------------- EXCEL HELPERS -------

def normalize_name(text):
    if not text:
        return ''
    d = unicodedata.normalize('NFD', str(text))
    return ''.join(c for c in d if unicodedata.category(c) != 'Mn').strip().lower()


HEADER_FILL = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
BORDER = Border(left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin'))

# Colores de cabecera por grupo de columnas, para distinguir de un vistazo Empresa/CNMC/Web/Bill.
# Paleta pensada para que combinen entre si (tonos "Office"): gris azulado, azul, verde,
# amarillo/dorado y naranja quemado para diferenciar Bill sin/con impuestos.
EMPRESA_FILL  = PatternFill(start_color="44546A", end_color="44546A", fill_type="solid")  # gris azulado
CNMC_FILL     = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")  # azul
WEB_FILL      = PatternFill(start_color="6B4C9A", end_color="6B4C9A", fill_type="solid")  # morado (antes verde: chocaba con el verde de "bajada")
BILL_FILL     = PatternFill(start_color="BF9000", end_color="BF9000", fill_type="solid")  # amarillo/dorado (sin impuestos)
BILL_IMP_FILL = PatternFill(start_color="C55A11", end_color="C55A11", fill_type="solid")  # naranja quemado (con impuestos)

# Para las comparativas antiguas (formato Energia/P1/P2, sin separacion CNMC/Web): un color por metrica.
ENERGIA_FILL = CNMC_FILL                                                                  # azul (reusa el mismo)
P1_FILL      = PatternFill(start_color="1B7A8C", end_color="1B7A8C", fill_type="solid")   # teal
P2_FILL      = PatternFill(start_color="B85C38", end_color="B85C38", fill_type="solid")   # terracota

DIFF_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")  # CNMC != Web (amarillo pastel)
UP_FILL   = PatternFill(start_color="FF5C5C", end_color="FF5C5C", fill_type="solid")  # subida vs semana anterior (rojo fuerte)
DOWN_FILL = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")  # bajada vs semana anterior (verde fuerte)

DIFF_FONT = Font(bold=True, color="000000")
UP_FONT   = Font(bold=True, color="FFFFFF")
DOWN_FONT = Font(bold=True, color="FFFFFF")

UMBRAL_DIFERENCIA = 0.05  # umbral para P1/P2 (EUR/kW/ano, decenas de euros) y para la Comparativa
UMBRAL_ENERGIA = 0.01     # umbral propio para Energia (EUR/kWh, escala 0.05-0.30): 0.05 ahi casi nunca se cumple


def marcar_si_difiere(ws, fila, col_cnmc, col_web, val_cnmc, val_web, umbral=UMBRAL_DIFERENCIA):
    """Resalta ambas celdas si CNMC y Web tienen dato y difieren mas del umbral."""
    if val_cnmc is not None and val_web is not None and abs(val_cnmc - val_web) > umbral:
        for c in (col_cnmc, col_web):
            cell = ws.cell(fila, c)
            cell.fill = DIFF_FILL
            cell.font = DIFF_FONT


def format_sheet(ws, num_cols, last_row, energy_cols=(2,), percent_cols=(), header_fills=None):
    header_fills = header_fills or {}
    for col in range(1, num_cols + 1):
        cell = ws.cell(1, col)
        cell.fill = header_fills.get(col, HEADER_FILL)
        cell.font = HEADER_FONT
        cell.border = BORDER
        cell.alignment = Alignment(horizontal='center', vertical='center')
    for row in range(2, last_row + 1):
        for col in range(1, num_cols + 1):
            cell = ws.cell(row, col)
            cell.border = BORDER
            cell.alignment = Alignment(horizontal='left' if col == 1 else 'right',
                                       vertical='center')
            if col >= 2:
                # el valor ya viene multiplicado por 100 (dif/prev*100), asi que se usa un
                # formato con el simbolo % literal en vez del '%' nativo de Excel (que
                # multiplicaria de nuevo por 100 al mostrarlo).
                if col in percent_cols:
                    cell.number_format = '0.00"%"'
                elif col in energy_cols:
                    cell.number_format = '0.0000'
                else:
                    cell.number_format = '0.00'
    ws.column_dimensions['A'].width = 22
    from openpyxl.utils import get_column_letter
    for col in range(2, num_cols + 1):
        ws.column_dimensions[get_column_letter(col)].width = 16


def read_sheet_records(ws):
    """Lee una pestana semanal (formato nuevo de 7 col o antiguo de 5) y devuelve
    {nombre_normalizado: {e_cnmc,p1_cnmc,p2_cnmc,e_web,p1_web,p2_web}} mapeando por
    el texto de la cabecera (robusto al cambio de formato)."""
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return {}
    headers = [str(h).lower() if h is not None else '' for h in rows[0]]
    colmap = {}
    for idx, h in enumerate(headers):
        if idx == 0 or 'fuente' in h:
            continue
        if not any(k in h for k in ('energ', 'p1', 'p2', 'precio')):
            continue
        src = 'web' if 'web' in h else 'cnmc'
        met = 'p1' if 'p1' in h else ('p2' if 'p2' in h else 'e')
        colmap[idx] = f'{met}_{src}'
    out = {}
    for row in rows[1:]:
        if not row or not row[0]:
            continue
        rec = {'e_cnmc': None, 'p1_cnmc': None, 'p2_cnmc': None,
               'e_web': None, 'p1_web': None, 'p2_web': None}
        for idx, key in colmap.items():
            if idx < len(row):
                rec[key] = row[idx]
        out[normalize_name(row[0])] = rec
    return out


# --------------------------------------------------------- EXTRACCION ----------

def main():
    print("=" * 70)
    print("CNMC BENCHMARK ENERGIA - DATOS REALES (API + WEB SCRAPING)")
    print("=" * 70 + "\n")

    hoy = datetime.now()
    fecha_hoy = hoy.strftime('%Y-%m-%d')
    # Pestana = lunes de la semana en curso (lunes-domingo), no la fecha exacta de hoy.
    # Asi, si se relanza el script varios dias dentro de la misma semana, se recrea/actualiza
    # la misma pestana en vez de acumular una por dia.
    lunes_semana = hoy - timedelta(days=hoy.weekday())
    pest_semana = lunes_semana.strftime('%Y-%m-%d')

    print(f"Fecha: {fecha_hoy} (semana del {pest_semana})")
    print(f"Extrayendo {len(TARGET_OFFERS)} empresas del API CNMC...\n")

    # SOLO ofertas de PRECIO FIJO: tipoElectricidad='TE'
    try:
        lista = cnmc_get('/ofertas/electricidad')
        todas = lista.get('resultadoComparador', []) or []
        ofertas = [o for o in todas if o.get('tipoElectricidad') == 'TE']
        n_unicas = sum(1 for o in ofertas if o.get('tienePrecioUnico') == 'S')
        print(f"  {len(todas)} ofertas ({len(ofertas)} de precio fijo, {n_unicas} de ellas precio unico)\n")
    except Exception as e:
        print(f"  ERROR al consultar el API: {e}\n")
        ofertas = []

    registros = []          # dicts: Empresa + CNMC (e/p1/p2) + Web (e/p1/p2)
    reg_by_name = {}
    for cfg in TARGET_OFFERS.values():
        reg = {'Empresa': cfg['brand'],
               'e_cnmc': None, 'p1_cnmc': None, 'p2_cnmc': None,
               'e_web': None, 'p1_web': None, 'p2_web': None}
        registros.append(reg)
        reg_by_name[cfg['brand']] = reg

    # ---- 1) CNMC: se intenta para TODAS las empresas (si tienen oferta de precio fijo) ----
    for legal, cfg in TARGET_OFFERS.items():
        marca = cfg['brand']
        search = cfg['search']
        reg = reg_by_name[marca]

        candidatas = [o for o in ofertas
                      if legal.upper() in (o.get('comercializadora', '') or '').upper()]
        if cfg.get('exclude'):
            candidatas = [o for o in candidatas
                          if not re.search(cfg['exclude'], (o.get('oferta', '') or '').lower())]

        if not candidatas:
            print(f"  --   {marca:22} (no en CNMC como precio fijo)")
            continue

        # REGLA: solo ofertas "Tarifa 1 Precio fijo" (tienePrecioUnico='S').
        # Si la empresa no tiene ninguna asi, NO se coge otra oferta -> queda sin CNMC.
        unicas = [o for o in candidatas if o.get('tienePrecioUnico') == 'S']
        if not unicas:
            print(f"  --   {marca:22} (sin oferta 'Precio fijo' en CNMC)")
            continue
        oferta = next((o for o in unicas if search.lower() in (o.get('oferta', '') or '').lower()), unicas[0])

        try:
            det = cnmc_get('/oferta', {'idOferta': oferta['id']})
            caract = det.get('caracteristicas', {})
            txt = caract.get('caracteristicas', '') if isinstance(caract, dict) else str(caract)
            pr = parse_prices(txt)

            # Endesa: energia CON descuento
            if marca == 'Endesa':
                mcd = re.search(r'con\s+\d+\s*%\s*descuento.{0,45}?([0-9][.,][0-9]{3,6})', txt, re.IGNORECASE)
                if mcd:
                    pr['energia'] = get_number(mcd.group(1))
            # Energya VM: energia CON descuento (tras "Promocion")
            if marca == 'Energya VM':
                mcd = re.search(r'Promoci.n[\s\S]{0,140}?([0-9][.,][0-9]{3,6})\s*[^0-9]{0,6}kWh', txt, re.IGNORECASE)
                if mcd:
                    pr['energia'] = get_number(mcd.group(1))

            reg['e_cnmc'] = pr['energia']
            reg['p1_cnmc'] = pr['p1']
            reg['p2_cnmc'] = pr['p2']

            if pr['energia'] is not None and pr['p1'] is not None and pr['p2'] is not None:
                print(f"  OK   {marca:22} E:{pr['energia']:.4f} | P1:{pr['p1']:.2f} | P2:{pr['p2']:.2f}")
            else:
                print(f"  WARN {marca:22} (parcial) E:{pr['energia']} P1:{pr['p1']} P2:{pr['p2']}")
        except Exception as e:
            print(f"  ERR  {marca:22} {e}")

    # ---- 2) WEB: se intenta SIEMPRE en todas las empresas (fuente principal). ----
    # Cada empresa con scraper propio afinado. Las que no tienen (webs con precio por
    # JavaScript o con varios productos ambiguos) no se pueden leer por HTTP -> quedan sin web.
    print("\nWeb scraping (SIEMPRE se intenta la web de cada empresa)...\n")
    for reg in registros:
        brand = reg['Empresa']
        scraper = SCRAPERS.get(brand)
        pr = scraper() if scraper else _none()
        reg['e_web'] = pr['energia']
        reg['p1_web'] = pr['p1']
        reg['p2_web'] = pr['p2']
        e_s = f"{pr['energia']:.4f}" if pr['energia'] is not None else "-"
        p1_s = f"{pr['p1']:.2f}" if pr['p1'] is not None else "-"
        p2_s = f"{pr['p2']:.2f}" if pr['p2'] is not None else "-"
        if pr['energia'] is not None or pr['p1'] is not None:
            print(f"  OK   {brand:22} E:{e_s:9} | P1:{p1_s:7} | P2:{p2_s}")
        elif scraper:
            print(f"  WARN {brand:22} (no extraido de web)")
        else:
            print(f"  --   {brand:22} (web no legible por HTTP: precio por JavaScript)")

    # ------------------------------------------------------------- EXCEL ------
    print("\nActualizando Excel...\n")
    CARPETA_SALIDA.mkdir(parents=True, exist_ok=True)
    ruta = CARPETA_SALIDA / NOMBRE_EXCEL

    if ruta.exists():
        wb = load_workbook(ruta)
    else:
        wb = Workbook()
        wb.remove(wb.active)

    # Leer la semana anterior mas reciente (solo para la comparativa)
    prev = sorted([s for s in wb.sheetnames
                   if re.fullmatch(r'\d{4}-\d{2}-\d{2}', s) and s != pest_semana],
                  reverse=True)
    datos_ant = {}
    ws_ant_name = None
    if prev:
        ws_ant_name = prev[0]
        datos_ant = read_sheet_records(wb[ws_ant_name])

    # Pestana semanal (recrear si existe): CNMC y Web en columnas separadas
    if pest_semana in wb.sheetnames:
        del wb[pest_semana]
    ws = wb.create_sheet(pest_semana, 0)
    headers = ['Empresa',
               'Energia CNMC (EUR/kWh)', 'P1 CNMC (EUR/kW/ano)', 'P2 CNMC (EUR/kW/ano)',
               'Energia Web (EUR/kWh)', 'P1 Web (EUR/kW/ano)', 'P2 Web (EUR/kW/ano)',
               'Bill Commodity (EUR/ano)', 'Bill Commodity con Impt. (EUR/ano)']
    ws.append(headers)
    for i, reg in enumerate(registros):
        fila = i + 2
        ws.append([reg['Empresa'],
                   reg['e_cnmc'], reg['p1_cnmc'], reg['p2_cnmc'],
                   reg['e_web'], reg['p1_web'], reg['p2_web'],
                   None, None])
        # Bill commodity = energia*Consumo + (P1+P2)*Potencia/12 -> se usa CNMC si esta
        # disponible; si no, Web. Si ninguna fuente tiene energia, "No disponible".
        ws.cell(fila, 8).value = (
            f'=IF(AND(B{fila}="",E{fila}=""),"No disponible",'
            f'IF(B{fila}<>"",B{fila},E{fila})*{CONSUMO_ANUAL_E}'
            f'+(IF(C{fila}<>"",C{fila},F{fila})+IF(D{fila}<>"",D{fila},G{fila}))*{POTENCIA}/12)'
        )
        # Bill commodity con impuestos = Bill commodity * (1+IEE) * (1+IVA). Sin descuentos.
        ws.cell(fila, 9).value = (
            f'=IF(ISNUMBER(H{fila}),H{fila}*(1+{IEE})*(1+{IVA}),"No disponible")'
        )
        # Resaltar cuando CNMC y Web difieren mas de UMBRAL_DIFERENCIA (Energia/P1/P2 de cada fuente).
        marcar_si_difiere(ws, fila, 2, 5, reg['e_cnmc'], reg['e_web'], umbral=UMBRAL_ENERGIA)
        marcar_si_difiere(ws, fila, 3, 6, reg['p1_cnmc'], reg['p1_web'])
        marcar_si_difiere(ws, fila, 4, 7, reg['p2_cnmc'], reg['p2_web'])
    format_sheet(ws, 9, len(registros) + 1, energy_cols=(2, 5), header_fills={
        1: EMPRESA_FILL,
        2: CNMC_FILL, 3: CNMC_FILL, 4: CNMC_FILL,
        5: WEB_FILL, 6: WEB_FILL, 7: WEB_FILL,
        8: BILL_FILL, 9: BILL_IMP_FILL,
    })
    print(f"  Pestana '{pest_semana}' creada")

    # Comparativa vs semana anterior (energia de cada fuente)
    if ws_ant_name:
        print(f"  Creando comparativa vs '{ws_ant_name}'...")
        nombre_comp = f"{pest_semana}-Comparativa"
        if nombre_comp in wb.sheetnames:
            del wb[nombre_comp]
        wsc = wb.create_sheet(nombre_comp, 1)
        wsc.append(['Empresa',
                    'E CNMC Actual', 'E CNMC Anterior', 'Dif (EUR/kWh)', 'Cambio %',
                    'E Web Actual', 'E Web Anterior', 'Dif (EUR/kWh)', 'Cambio %'])
        for i, reg in enumerate(registros):
            fila_n = i + 2
            ant = datos_ant.get(normalize_name(reg['Empresa']), {})
            fila = [reg['Empresa']]
            difs = []
            for act, prev_v in ((reg['e_cnmc'], ant.get('e_cnmc')),
                                (reg['e_web'], ant.get('e_web'))):
                dif = pct = None
                if act is not None and prev_v not in (None, 0):
                    dif = act - prev_v
                    pct = dif / prev_v * 100
                fila += [act, prev_v, dif, pct]
                difs.append(dif)
            wsc.append(fila)
            # Resaltar subida (rojo) / bajada (verde) vs la semana anterior, si el cambio
            # supera UMBRAL_ENERGIA (aqui todo son valores de Energia, EUR/kWh), en Dif y Cambio %.
            for dif, cols in zip(difs, ((4, 5), (8, 9))):
                if dif is not None and abs(dif) > UMBRAL_ENERGIA:
                    fill = UP_FILL if dif > 0 else DOWN_FILL
                    font = UP_FONT if dif > 0 else DOWN_FONT
                    for c in cols:
                        cell = wsc.cell(fila_n, c)
                        cell.fill = fill
                        cell.font = font
        format_sheet(wsc, 9, len(registros) + 1, energy_cols=(2, 3, 4, 6, 7, 8), percent_cols=(5, 9), header_fills={
            1: EMPRESA_FILL,
            2: CNMC_FILL, 3: CNMC_FILL, 4: CNMC_FILL, 5: CNMC_FILL,
            6: WEB_FILL, 7: WEB_FILL, 8: WEB_FILL, 9: WEB_FILL,
        })
        print("  Comparativa creada")

    wb.save(ruta)
    print(f"\n  Pestanas: {', '.join(wb.sheetnames)}")

    con_cnmc = sum(1 for r in registros if r['e_cnmc'] is not None)
    con_web = sum(1 for r in registros if r['e_web'] is not None)
    con_any = sum(1 for r in registros if r['e_cnmc'] is not None or r['e_web'] is not None)
    print(f"\nArchivo guardado en: {ruta}")
    print(f"Total empresas: {len(registros)}  |  con CNMC: {con_cnmc}  |  con Web: {con_web}  |  con algun dato: {con_any}")


if __name__ == '__main__':
    main()
